import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

class ExportManager:
    @staticmethod
    def export_to_excel(data_dicts, file_path):
        """
        Exporta uma lista de dicionários para um arquivo Excel com formatação.
        """
        df = pd.DataFrame(data_dicts)
        
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='HistoricoEstoque')
        workbook = writer.book
        worksheet = writer.sheets['HistoricoEstoque']

        ExportManager._apply_excel_formatting(worksheet)

        writer.close()

    @staticmethod
    def _apply_excel_formatting(worksheet):
        HEADER_FILL_COLOR = "00008B" # Azul escuro
        HEADER_FONT_COLOR = "FFFFFF" # Branco
        header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
        header_font = Font(color=HEADER_FONT_COLOR, bold=True)

        for col_num, cell in enumerate(worksheet[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

            for cell in column_cells:
                cell.alignment = Alignment(horizontal='center', vertical='center')
